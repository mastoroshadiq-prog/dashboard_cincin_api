"""
Extract production data dengan formula yang benar:
- Luas di col 12
- Real Ton: 2023=col 153, 2024=col 162, 2025=col 171
- Poten Ton: 2023=col 156, 2024=col 165, 2025=col 174
- Ton/Ha = Ton / Luas
"""

import openpyxl
import json

wb = openpyxl.load_workbook(r'f:\PythonProjects\poac_cincin_api\dashboard-cincin-api\poac_sim\data\input\data_gabungan.xlsx', data_only=True)
sheet = wb.active

# Extract all blocks
all_blocks = {}

for row in range(7, 700):
    block_code = sheet.cell(row=row, column=1).value
    division = sheet.cell(row=row, column=6).value
    
    if not block_code or not division:
        continue
    if str(block_code).upper() in ['BLOK', 'BLOCK', 'ESTATE', 'DIVISI']:
        continue
    
    # Luas from col 12
    try:
        luas = float(sheet.cell(row=row, column=12).value or 0)
    except:
        luas = 0
    
    if luas <= 0:
        continue
    
    # Production Ton (total)
    try:
        real_ton_2023 = float(sheet.cell(row=row, column=153).value or 0)
        poten_ton_2023 = float(sheet.cell(row=row, column=156).value or 0)
        real_ton_2024 = float(sheet.cell(row=row, column=162).value or 0)
        poten_ton_2024 = float(sheet.cell(row=row, column=165).value or 0)
        real_ton_2025 = float(sheet.cell(row=row, column=171).value or 0)
        poten_ton_2025 = float(sheet.cell(row=row, column=174).value or 0)
    except:
        continue
    
    # Convert to Ton/Ha
    real_tha_2023 = real_ton_2023 / luas
    poten_tha_2023 = poten_ton_2023 / luas
    real_tha_2024 = real_ton_2024 / luas
    poten_tha_2024 = poten_ton_2024 / luas
    real_tha_2025 = real_ton_2025 / luas
    poten_tha_2025 = poten_ton_2025 / luas
    
    # Calculate trend
    if real_tha_2023 > 0:
        change_pct = ((real_tha_2025 - real_tha_2023) / real_tha_2023) * 100
    else:
        change_pct = 0
    
    all_blocks[block_code] = {
        'block_code': block_code,
        'division': division,
        'luas_ha': round(luas, 2),
        'yields': {
            '2023': {
                'real_ton_ha': round(real_tha_2023, 2),
                'poten_ton_ha': round(poten_tha_2023, 2),
                'gap_pct': round(((poten_tha_2023 - real_tha_2023) / poten_tha_2023 * 100) if poten_tha_2023 > 0 else 0, 1)
            },
            '2024': {
                'real_ton_ha': round(real_tha_2024, 2),
                'poten_ton_ha': round(poten_tha_2024, 2),
                'gap_pct': round(((poten_tha_2024 - real_tha_2024) / poten_tha_2024 * 100) if poten_tha_2024 > 0 else 0, 1)
            },
            '2025': {
                'real_ton_ha': round(real_tha_2025, 2),
                'poten_ton_ha': round(poten_tha_2025, 2),
                'gap_pct': round(((poten_tha_2025 - real_tha_2025) / poten_tha_2025 * 100) if poten_tha_2025 > 0 else 0, 1)
            }
        },
        'change_2023_2025_pct': round(change_pct, 1),
        'trend': 'DECLINING' if change_pct < -5 else ('INCREASING' if change_pct > 5 else 'STABLE')
    }

print(f"Total blocks extracted: {len(all_blocks)}")

# Show AME02 summary
ame02 = {k: v for k, v in all_blocks.items() if v['division'] == 'AME02'}
print(f"\n=== AME02 ({len(ame02)} blocks) ===")

dec = [k for k, v in ame02.items() if v['trend'] == 'DECLINING']
stb = [k for k, v in ame02.items() if v['trend'] == 'STABLE']
inc = [k for k, v in ame02.items() if v['trend'] == 'INCREASING']

print(f"DECLINING: {len(dec)}")
print(f"STABLE: {len(stb)}")
print(f"INCREASING: {len(inc)}")

# Show sample
print("\n--- Sample Declining ---")
for k in sorted(dec, key=lambda x: ame02[x]['change_2023_2025_pct'])[:5]:
    v = ame02[k]
    print(f"  {k}: {v['change_2023_2025_pct']:+.1f}% | 2023: {v['yields']['2023']['real_ton_ha']:.1f} -> 2025: {v['yields']['2025']['real_ton_ha']:.1f} T/Ha")

print("\n--- Sample Increasing ---")
for k in sorted(inc, key=lambda x: -ame02[x]['change_2023_2025_pct'])[:5]:
    v = ame02[k]
    print(f"  {k}: {v['change_2023_2025_pct']:+.1f}% | 2023: {v['yields']['2023']['real_ton_ha']:.1f} -> 2025: {v['yields']['2025']['real_ton_ha']:.1f} T/Ha")

# Save corrected data
with open('complete_historical_yields.json', 'w') as f:
    json.dump(all_blocks, f, indent=2)
print("\n✅ Saved corrected data to complete_historical_yields.json")
