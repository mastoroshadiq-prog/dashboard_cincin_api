"""
Analisis data_gabungan.xlsx untuk menemukan:
1. Blok dengan tren PENURUNAN produksi (2023-2025)
2. Blok dengan tren KENAIKAN produksi (2023-2025)
3. Blok dengan tren STABIL
Fokus pada divisi AME02
"""

import openpyxl
from collections import defaultdict

# Open the Excel file
wb = openpyxl.load_workbook(r'f:\PythonProjects\poac_cincin_api\dashboard-cincin-api\poac_sim\data\input\data_gabungan.xlsx', data_only=True)
sheet = wb.active

# First, let's understand the column structure for production data
# From previous analysis:
# - Col 1: Block code
# - Col 6: Division
# - Col 151+: 2023 production data
# - Col 160+: 2024 production data  
# - Col 169+: 2025 production data

# Let's find the exact columns for Realisasi (Ton/Ha) for each year
print("=== FINDING COLUMN STRUCTURE FOR PRODUCTION DATA ===")
for col in range(145, 180):
    r4 = sheet.cell(row=4, column=col).value or ''
    r5 = sheet.cell(row=5, column=col).value or ''
    r6 = sheet.cell(row=6, column=col).value or ''
    if r4 or r5 or r6:
        if 'Ton' in str(r6) or 'Real' in str(r5) or '2023' in str(r4) or '2024' in str(r4) or '2025' in str(r4):
            print(f"Col {col}: R4='{r4}' | R5='{r5}' | R6='{r6}'")

# Let's also check the basic block info columns
print("\n=== BLOCK INFO COLUMNS ===")
for col in range(1, 10):
    r4 = sheet.cell(row=4, column=col).value or ''
    r5 = sheet.cell(row=5, column=col).value or ''
    r6 = sheet.cell(row=6, column=col).value or ''
    print(f"Col {col}: R4='{r4}' | R5='{r5}' | R6='{r6}'")

# Check sample data row (row 10)
print("\n=== SAMPLE DATA ROW (Row 10) ===")
sample_cols = [1, 2, 4, 6, 153, 156, 162, 165, 171, 174]
for col in sample_cols:
    val = sheet.cell(row=10, column=col).value
    print(f"Col {col}: {val}")
