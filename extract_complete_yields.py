"""
Ekstrak data produksi lengkap dari data_gabungan.xlsx
dan generate COMPLETE_HISTORICAL_YIELDS untuk semua blok
"""

import openpyxl
import json

# Open the Excel file
wb = openpyxl.load_workbook(r'f:\PythonProjects\poac_cincin_api\dashboard-cincin-api\poac_sim\data\input\data_gabungan.xlsx', data_only=True)
sheet = wb.active

# Column mapping based on analysis:
# Col 1: Block code
# Col 6: Division  
# Col 10: Luas (Ha)
# Production columns (Ton total, need to divide by Luas for Ton/Ha):
# Col 153: 2023 Real Ton
# Col 156: 2023 Poten Ton
# Col 162: 2024 Real Ton  
# Col 165: 2024 Poten Ton
# Col 171: 2025 Real Ton
# Col 174: 2025 Poten Ton

# First, find the Luas column
print("=== FINDING LUAS COLUMN ===")
for col in range(8, 15):
    r4 = sheet.cell(row=4, column=col).value or ''
    r5 = sheet.cell(row=5, column=col).value or ''
    r6 = sheet.cell(row=6, column=col).value or ''
    print(f"Col {col}: R4='{r4}' | R5='{r5}' | R6='{r6}'")

# Check row 10 for sample
print("\n=== SAMPLE ROW 10 ===")
for col in [1, 6, 10, 11, 12, 153, 156, 162, 165, 171, 174]:
    val = sheet.cell(row=10, column=col).value
    print(f"Col {col}: {val}")

# Extract all blocks data
all_blocks = {}
division_counts = {}

# Start from row 7 (data rows)
for row in range(7, 700):
    block_code = sheet.cell(row=row, column=1).value
    division = sheet.cell(row=row, column=6).value
    
    if not block_code or not division:
        continue
    
    # Skip header rows
    if block_code in ['BLOK', 'Block', 'ESTATE'] or 'DIVISI' in str(block_code):
        continue
        
    # Get Luas - try column 10, 11, 12
    luas = sheet.cell(row=row, column=10).value or sheet.cell(row=row, column=11).value or 0
    try:
        luas = float(luas) if luas else 0
    except:
        luas = 0
    
    # Get production data
    try:
        real_2023 = float(sheet.cell(row=row, column=153).value or 0)
        poten_2023 = float(sheet.cell(row=row, column=156).value or 0)
        real_2024 = float(sheet.cell(row=row, column=162).value or 0)
        poten_2024 = float(sheet.cell(row=row, column=165).value or 0)
        real_2025 = float(sheet.cell(row=row, column=171).value or 0)
        poten_2025 = float(sheet.cell(row=row, column=174).value or 0)
    except:
        continue
    
    # Calculate Ton/Ha (if luas > 0)
    if luas > 0:
        real_tha_2023 = real_2023 / luas
        poten_tha_2023 = poten_2023 / luas
        real_tha_2024 = real_2024 / luas
        poten_tha_2024 = poten_2024 / luas
        real_tha_2025 = real_2025 / luas
        poten_tha_2025 = poten_2025 / luas
    else:
        real_tha_2023 = real_tha_2024 = real_tha_2025 = 0
        poten_tha_2023 = poten_tha_2024 = poten_tha_2025 = 0
    
    # Calculate trend
    if real_tha_2023 > 0:
        change_pct = ((real_tha_2025 - real_tha_2023) / real_tha_2023) * 100
    else:
        change_pct = 0
    
    # Classify trend
    if change_pct < -5:
        trend = 'DECLINING'
    elif change_pct > 5:
        trend = 'INCREASING'
    else:
        trend = 'STABLE'
    
    all_blocks[block_code] = {
        'block_code': block_code,
        'division': division,
        'luas_ha': round(luas, 2),
        'yields': {
            2023: {
                'real_ton_ha': round(real_tha_2023, 2),
                'poten_ton_ha': round(poten_tha_2023, 2),
                'gap_pct': round(((poten_tha_2023 - real_tha_2023) / poten_tha_2023 * 100) if poten_tha_2023 > 0 else 0, 1)
            },
            2024: {
                'real_ton_ha': round(real_tha_2024, 2),
                'poten_ton_ha': round(poten_tha_2024, 2),
                'gap_pct': round(((poten_tha_2024 - real_tha_2024) / poten_tha_2024 * 100) if poten_tha_2024 > 0 else 0, 1)
            },
            2025: {
                'real_ton_ha': round(real_tha_2025, 2),
                'poten_ton_ha': round(poten_tha_2025, 2),
                'gap_pct': round(((poten_tha_2025 - real_tha_2025) / poten_tha_2025 * 100) if poten_tha_2025 > 0 else 0, 1)
            }
        },
        'change_2023_2025_pct': round(change_pct, 1),
        'trend': trend
    }
    
    # Count by division
    if division not in division_counts:
        division_counts[division] = {'total': 0, 'declining': 0, 'stable': 0, 'increasing': 0}
    division_counts[division]['total'] += 1
    division_counts[division][trend.lower()] += 1

print(f"\n=== TOTAL BLOCKS EXTRACTED: {len(all_blocks)} ===")

print("\n=== DIVISION SUMMARY ===")
for div, counts in sorted(division_counts.items()):
    print(f"{div}: {counts['total']} blocks | Declining: {counts['declining']} | Stable: {counts['stable']} | Increasing: {counts['increasing']}")

# Show AME02 blocks detail
print("\n=== AME02 BLOCKS DETAIL ===")
ame02_blocks = {k: v for k, v in all_blocks.items() if v['division'] == 'AME02'}
print(f"Total AME02 blocks: {len(ame02_blocks)}")

# Sort by trend
declining = [(k, v) for k, v in ame02_blocks.items() if v['trend'] == 'DECLINING']
increasing = [(k, v) for k, v in ame02_blocks.items() if v['trend'] == 'INCREASING']
stable = [(k, v) for k, v in ame02_blocks.items() if v['trend'] == 'STABLE']

print(f"\nDECLINING ({len(declining)}):")
for code, data in sorted(declining, key=lambda x: x[1]['change_2023_2025_pct'])[:10]:
    print(f"  {code}: {data['change_2023_2025_pct']:+.1f}% | 2023: {data['yields'][2023]['real_ton_ha']:.1f} -> 2025: {data['yields'][2025]['real_ton_ha']:.1f} T/Ha")

print(f"\nINCREASING ({len(increasing)}):")
for code, data in sorted(increasing, key=lambda x: -x[1]['change_2023_2025_pct'])[:10]:
    print(f"  {code}: {data['change_2023_2025_pct']:+.1f}% | 2023: {data['yields'][2023]['real_ton_ha']:.1f} -> 2025: {data['yields'][2025]['real_ton_ha']:.1f} T/Ha")

print(f"\nSTABLE ({len(stable)}):")
for code, data in stable[:5]:
    print(f"  {code}: {data['change_2023_2025_pct']:+.1f}% | 2023: {data['yields'][2023]['real_ton_ha']:.1f} -> 2025: {data['yields'][2025]['real_ton_ha']:.1f} T/Ha")

# Save to JSON for later use
with open('complete_historical_yields.json', 'w') as f:
    json.dump(all_blocks, f, indent=2)
print("\n✅ Data saved to complete_historical_yields.json")
