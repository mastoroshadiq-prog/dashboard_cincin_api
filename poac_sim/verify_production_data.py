import pandas as pd
import openpyxl

# Open the Excel file
wb = openpyxl.load_workbook('data/input/Realisasi vs Potensi PT SR.xlsx', data_only=True)
ws = wb.active

print("="*80)
print("🔍 VERIFIKASI DATA PRODUKSI TAHUN 2025 - BLOK D006A & D007A")
print("="*80)

# Find D006A and D007A rows
d006a_row = 99  # From previous output
d007a_row = 100

# Check header structure - look at rows 1-10 for all columns
print("\n📋 Struktur Header File (Rows 1-10, Columns 1-30):\n")

for row_idx in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
    row_data = []
    for col_idx in range(1, 31):
        val = ws.cell(row=row_idx, column=col_idx).value
        if val:
            row_data.append(f"C{col_idx}:{val}")
    if row_data:
        print(f"Row {row_idx}: {' | '.join(row_data[:10])}")  # Show first 10 non-empty

# Based on the file structure from session summary, let's check specific columns
# The file likely has: Blok (col 3), Ha (col 4), and production data in later columns

print("\n" + "="*80)
print("📊 DATA LENGKAP BLOK D006A (Row 99)")
print("="*80)

print(f"\nBlok: {ws.cell(row=d006a_row, column=3).value}")
print(f"Luas (Ha): {ws.cell(row=d006a_row, column=4).value}")
print(f"Tahun Tanam: {ws.cell(row=d006a_row, column=7).value}")

# Check columns that might contain 2025 data
# Based on output, columns around 109-113 seem relevant
print("\n🔍 Mencari kolom Potensi dan Realisasi 2025...")
print("\nSample data dari berbagai kolom:")

# Check columns 1-20 first
for col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
    val = ws.cell(row=d006a_row, column=col_idx).value
    header = ws.cell(row=8, column=col_idx).value  # Try row 8 for header
    print(f"  Col {col_idx} (Header: {header}): {val}")

print("\n" + "="*80)
print("📊 DATA LENGKAP BLOK D007A (Row 100)")
print("="*80)

print(f"\nBlok: {ws.cell(row=d007a_row, column=3).value}")
print(f"Luas (Ha): {ws.cell(row=d007a_row, column=4).value}")
print(f"Tahun Tanam: {ws.cell(row=d007a_row, column=7).value}")

print("\nSample data dari berbagai kolom:")
for col_idx in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]:
    val = ws.cell(row=d007a_row, column=col_idx).value
    header = ws.cell(row=8, column=col_idx).value
    print(f"  Col {col_idx} (Header: {header}): {val}")

# Now let's specifically look for columns with values matching our dashboard
# Dashboard shows: D006A Potensi=17.30, Realisasi=0.79
# Dashboard shows: D007A Potensi=17.53, Realisasi=0.30

print("\n" + "="*80)
print("🎯 MENCARI KOLOM YANG COCOK DENGAN NILAI DASHBOARD")
print("="*80)
print("\nTarget nilai yang dicari:")
print("  D006A: Potensi ≈ 17.30, Realisasi ≈ 0.79")
print("  D007A: Potensi ≈ 17.53, Realisasi ≈ 0.30")

print("\n🔍 Scanning semua kolom...")

matches_d006a = []
matches_d007a = []

for col_idx in range(1, min(ws.max_column + 1, 120)):
    val_d006a = ws.cell(row=d006a_row, column=col_idx).value
    val_d007a = ws.cell(row=d007a_row, column=col_idx).value
    
    # Check for Potensi values (around 17)
    if val_d006a and isinstance(val_d006a, (int, float)):
        if 17 <= val_d006a <= 18:
            matches_d006a.append((col_idx, val_d006a, 'Potensi?'))
        elif 0.7 <= val_d006a <= 0.9:
            matches_d006a.append((col_idx, val_d006a, 'Realisasi?'))
    
    if val_d007a and isinstance(val_d007a, (int, float)):
        if 17 <= val_d007a <= 18:
            matches_d007a.append((col_idx, val_d007a, 'Potensi?'))
        elif 0.2 <= val_d007a <= 0.4:
            matches_d007a.append((col_idx, val_d007a, 'Realisasi?'))

print("\n✅ HASIL PENCARIAN D006A:")
for col_idx, val, label in matches_d006a:
    header = ws.cell(row=8, column=col_idx).value
    print(f"  Col {col_idx} (Header: {header}): {val} → {label}")

print("\n✅ HASIL PENCARIAN D007A:")
for col_idx, val, label in matches_d007a:
    header = ws.cell(row=8, column=col_idx).value
    print(f"  Col {col_idx} (Header: {header}): {val} → {label}")

print("\n" + "="*80)
print("✅ VERIFIKASI SELESAI")
print("="*80)
