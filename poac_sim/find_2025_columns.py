import pandas as pd
import openpyxl

# Open Excel file
wb = openpyxl.load_workbook('poac_sim/data/input/data_gabungan.xlsx', data_only=True)
ws = wb.active

print('='*80)
print('🔍 MENCARI DATA PRODUKSI 2025 - BERDASARKAN SCREENSHOT')
print('='*80)

# From screenshot, we know:
# - Row 1: "2024" and "2025" headers
# - Row 2: "Real", "Potensi", "Real vs Potensi"
# - Row 3: "BJR Kg", "Jum Jlg", "Ton"

# Let's scan row 1 for "2025"
print('\n📋 Scanning Row 1 for "2025":')
for col in range(1, min(ws.max_column + 1, 150)):
    val = ws.cell(row=1, column=col).value
    if val and '2025' in str(val):
        print(f'  Found "2025" at column {col}')
        
        # Check surrounding columns
        print(f'\n  Checking columns around {col}:')
        for c in range(max(1, col-5), min(ws.max_column + 1, col+15)):
            r1 = ws.cell(row=1, column=c).value
            r2 = ws.cell(row=2, column=c).value
            r3 = ws.cell(row=3, column=c).value
            r4 = ws.cell(row=4, column=c).value
            print(f'    Col {c}: R1={r1}, R2={r2}, R3={r3}, R4={r4}')
        break

# Find F008A row
print('\n📊 Finding F008A and D001A rows:')
f008a_row = None
d001a_row = None

for row in range(1, min(ws.max_row + 1, 200)):
    val = ws.cell(row=row, column=1).value
    if val == 'F008A':
        f008a_row = row
        print(f'  F008A found at row {f008a_row}')
    elif val == 'D001A':
        d001a_row = row
        print(f'  D001A found at row {d001a_row}')
    
    if f008a_row and d001a_row:
        break

# Once we find the 2025 columns, extract data
print('\n' + '='*80)
print('📊 EKSTRAKSI DATA PRODUKSI 2025')
print('='*80)

# Based on typical structure, if 2025 starts at a certain column,
# Real should be first 3 columns (BJR Kg, Jum Jlg, Ton)
# Potensi should be next 3 columns
# Real vs Potensi should be next 3 columns

# Let's manually check columns that might contain 2025 data
# From previous output, we saw data in columns 107-115
# Let's check if there's a pattern

if f008a_row:
    print('\nF008A data inspection:')
    print('  Luas:', ws.cell(row=f008a_row, column=13).value, 'Ha')
    
    # Check columns 100-120 for Ton values
    print('\n  Columns 100-120 (looking for Ton values 10-30):')
    for col in range(100, 121):
        val = ws.cell(row=f008a_row, column=col).value
        if val and isinstance(val, (int, float)):
            if 10 < val < 30:  # Reasonable Ton/Ha range
                r1 = ws.cell(row=1, column=col).value
                r2 = ws.cell(row=2, column=col).value
                r3 = ws.cell(row=3, column=col).value
                print(f'    Col {col}: {val:.2f} (R1={r1}, R2={r2}, R3={r3})')

print('\n✅ Selesai!')
print('\nCATATAN: Berdasarkan hasil di atas, identifikasi kolom mana yang:')
print('  - 2025 Real Ton')
print('  - 2025 Potensi Ton')
print('  - 2025 Gap Ton')
