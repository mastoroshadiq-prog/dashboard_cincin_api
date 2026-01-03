"""
SYSTEMATIC EXTRACTION - Find Potensi/Realisasi for D001A and F008A
User confirmed data IS in data_gabungan.xlsx
Be patient and systematic!
"""
from openpyxl import load_workbook

wb = load_workbook('poac_sim/data/input/data_gabungan.xlsx', data_only=True)
ws = wb.active

print("🔍 SYSTEMATIC SEARCH FOR PRODUCTION DATA")
print("="*70)

# Known facts:
# D001A: Row 88, Luas=25.8 Ha (col 12), Potensi=22.13 ton/ha, Realisasi=17.42 ton/ha
# F008A: Row 94, Potensi=19.52 ton/ha, Realisasi=21.22 ton/ha

blocks = {
    'D001A': {
        'row': 88,
        'luas_ha': 25.8,
        'potensi_ton_ha': 22.13,
        'realisasi_ton_ha': 17.42,
        'gap_ton_ha': -4.71
    },
    'F008A': {
        'row': 94,
        'luas_ha': 29.6,  # From dashboard
        'potensi_ton_ha': 19.52,
        'realisasi_ton_ha': 21.22,
        'gap_ton_ha': 1.71
    }
}

for block_code, info in blocks.items():
    print(f"\n{'='*70}")
    print(f"{block_code} (Row {info['row']})")
    print(f"{'='*70}")
    
    row = info['row']
    luas = info['luas_ha']
    
    # Calculate expected total tons
    expected_real_total = info['realisasi_ton_ha'] * luas
    expected_pot_total = info['potensi_ton_ha'] * luas
    expected_gap_total = info['gap_ton_ha'] * luas
    
    print(f"\nExpected total tons:")
    print(f"  Realisasi: {expected_real_total:.1f} ton")
    print(f"  Potensi: {expected_pot_total:.1f} ton")
    print(f"  Gap: {expected_gap_total:.1f} ton")
    
    print(f"\nSearching columns 1-177 for matches...")
    
    matches = []
    for col in range(1, 178):
        val = ws.cell(row, col).value
        if val is None:
            continue
            
        try:
            num_val = float(val)
            
            # Check if close to expected values (within 5%)
            if abs(num_val - expected_real_total) / expected_real_total < 0.05:
                matches.append((col, num_val, 'REALISASI TOTAL'))
                print(f"  ✅ Col {col}: {num_val:.2f} ← REALISASI TOTAL MATCH!")
                
            elif abs(num_val - expected_pot_total) / expected_pot_total < 0.05:
                matches.append((col, num_val, 'POTENSI TOTAL'))
                print(f"  ✅ Col {col}: {num_val:.2f} ← POTENSI TOTAL MATCH!")
                
            elif abs(num_val - info['realisasi_ton_ha']) < 0.1:
                matches.append((col, num_val, 'REALISASI TON/HA'))
                print(f"  ✅ Col {col}: {num_val:.2f} ← REALISASI TON/HA MATCH!")
                
            elif abs(num_val - info['potensi_ton_ha']) < 0.1:
                matches.append((col, num_val, 'POTENSI TON/HA'))
                print(f"  ✅ Col {col}: {num_val:.2f} ← POTENSI TON/HA MATCH!")
                
        except (ValueError, TypeError):
            pass
    
    if not matches:
        print("  ❌ No matches found!")
        print("\n  Showing ALL numeric columns 100-130:")
        for col in range(100, 131):
            val = ws.cell(row, col).value
            if val not in [0, None, '0']:
                print(f"    Col {col}: {val}")

wb.close()

print("\n" + "="*70)
print("CONCLUSION:")
print("If matches found → those are the column indices to use")
print("If no matches → data might be in different sheet or calculated differently")
