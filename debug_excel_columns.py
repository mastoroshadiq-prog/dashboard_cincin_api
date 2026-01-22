"""
Debug dan perbaiki ekstraksi data dari data_gabungan.xlsx
Fokus pada menemukan kolom yang benar untuk Luas dan Produksi Ton/Ha
"""

import openpyxl

# Open the Excel file
wb = openpyxl.load_workbook(r'f:\PythonProjects\poac_cincin_api\dashboard-cincin-api\poac_sim\data\input\data_gabungan.xlsx', data_only=True)
sheet = wb.active

# Check header rows more carefully
print("=== DETAILED HEADER ANALYSIS ===")
print("\nRow 4-6 for cols 145-180:")
for col in range(145, 180):
    r3 = sheet.cell(row=3, column=col).value or ''
    r4 = sheet.cell(row=4, column=col).value or ''
    r5 = sheet.cell(row=5, column=col).value or ''
    r6 = sheet.cell(row=6, column=col).value or ''
    if any([r3, r4, r5, r6]):
        print(f"Col {col}: R3='{r3}' | R4='{r4}' | R5='{r5}' | R6='{r6}'")

# Check specific data row for AME02 block
print("\n=== SAMPLE DATA ROW 10 (D001A from AME02) ===")
print(f"Col 1 (Block): {sheet.cell(row=10, column=1).value}")
print(f"Col 6 (Division): {sheet.cell(row=10, column=6).value}")

# Check Luas columns more carefully
for col in [10, 11, 12, 13, 14]:
    val = sheet.cell(row=10, column=col).value
    header3 = sheet.cell(row=3, column=col).value or ''
    header4 = sheet.cell(row=4, column=col).value or ''
    header5 = sheet.cell(row=5, column=col).value or ''
    header6 = sheet.cell(row=6, column=col).value or ''
    print(f"Col {col}: Value={val} | Headers: R3='{header3}' R4='{header4}' R5='{header5}' R6='{header6}'")

# Check production columns - look for Ton/Ha not just Ton
print("\n=== LOOKING FOR TON/HA COLUMNS ===")
for col in range(145, 180):
    r5 = str(sheet.cell(row=5, column=col).value or '')
    r6 = str(sheet.cell(row=6, column=col).value or '')
    if 'Ton' in r6 or 'Ha' in r6:
        val = sheet.cell(row=10, column=col).value
        print(f"Col {col}: R5='{r5}' | R6='{r6}' | Sample value={val}")

# The issue might be that we're using Ton total not Ton/Ha
# Let's check columns around Real and Potensi
print("\n=== CHECKING REAL/POTENSI COLUMNS ===")
for col in range(150, 180):
    r4 = str(sheet.cell(row=4, column=col).value or '')
    r5 = str(sheet.cell(row=5, column=col).value or '')
    r6 = str(sheet.cell(row=6, column=col).value or '')
    if 'Real' in r5 or 'Poten' in r5:
        val = sheet.cell(row=10, column=col).value
        print(f"Col {col}: R4='{r4}' | R5='{r5}' | R6='{r6}' | Sample value={val}")

# Check if there are Ton/Ha columns directly
print("\n=== LOOKING FOR DIRECT TON/HA ===")
for col in range(1, 200):
    r6 = str(sheet.cell(row=6, column=col).value or '').lower()
    if 'ton/ha' in r6 or 't/ha' in r6:
        val = sheet.cell(row=10, column=col).value
        r4 = sheet.cell(row=4, column=col).value or ''
        r5 = sheet.cell(row=5, column=col).value or ''
        print(f"Col {col}: R4='{r4}' | R5='{r5}' | R6='{r6}' | Sample value={val}")
