"""
Find Ton/Ha columns in data_gabungan.xlsx
"""

import openpyxl

wb = openpyxl.load_workbook(r'f:\PythonProjects\poac_cincin_api\dashboard-cincin-api\poac_sim\data\input\data_gabungan.xlsx', data_only=True)
sheet = wb.active

# Look for Ton/Ha in row 6
print("=== COLUMNS WITH TON/HA ===")
for col in range(1, 200):
    r6 = str(sheet.cell(row=6, column=col).value or '')
    if 'Ton' in r6 and 'Ha' in r6:
        r4 = sheet.cell(row=4, column=col).value or ''
        r5 = sheet.cell(row=5, column=col).value or ''
        val = sheet.cell(row=10, column=col).value
        print(f"Col {col}: R4='{r4}' R5='{r5}' R6='{r6}' | Value={val}")

# Alternative: look for T/Ha
print("\n=== COLUMNS WITH T/Ha ===")
for col in range(1, 200):
    r6 = str(sheet.cell(row=6, column=col).value or '')
    if 'T/Ha' in r6 or 't/ha' in r6.lower():
        r4 = sheet.cell(row=4, column=col).value or ''
        r5 = sheet.cell(row=5, column=col).value or ''
        val = sheet.cell(row=10, column=col).value
        print(f"Col {col}: R4='{r4}' R5='{r5}' R6='{r6}' | Value={val}")

# Check for Luas column
print("\n=== LUAS COLUMN ===")
for col in range(1, 50):
    r4 = str(sheet.cell(row=4, column=col).value or '')
    r5 = str(sheet.cell(row=5, column=col).value or '')
    r6 = str(sheet.cell(row=6, column=col).value or '')
    if 'LUAS' in r4.upper() or 'LUAS' in r5.upper() or 'HA' in r4.upper():
        val = sheet.cell(row=10, column=col).value
        print(f"Col {col}: R4='{r4}' R5='{r5}' R6='{r6}' | Value={val}")

# Check row 10 for columns 150-155 (around 2023 Real)
print("\n=== 2023 REAL AREA COLS 150-160 ===")
for col in range(150, 165):
    r4 = str(sheet.cell(row=4, column=col).value or '')
    r5 = str(sheet.cell(row=5, column=col).value or '')
    r6 = str(sheet.cell(row=6, column=col).value or '')
    val = sheet.cell(row=10, column=col).value
    print(f"Col {col}: R4='{r4}' R5='{r5}' R6='{r6}' | Value={val}")
