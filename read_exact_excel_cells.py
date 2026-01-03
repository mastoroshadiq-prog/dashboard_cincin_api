"""
Use openpyxl to read EXACT cell references P109, P112 (for F008A row 109)
"""
from openpyxl import load_workbook
import json

print("📍 READING EXACT EXCEL CELLS")
print("="*70)

wb = load_workbook('poac_sim/data/input/data_gabungan.xlsx', data_only=True)
ws = wb.active

print(f"✅ Loaded workbook, active sheet: {ws.title}")
print(f"Dimensions: {ws.max_row} rows × {ws.max_column} columns")

# Find F008A which user said is around row 109
print("\n🔎 Searching for F008A...")

f008_row = None
for row_idx in range(1, min(200, ws.max_row + 1)):
    cell_val = ws.cell(row_idx, 1).value  # Column A (index 1)
    if cell_val == 'F008A':
        f008_row = row_idx
        print(f"✅ Found F008A at row {row_idx}")
        break

if f008_row:
    # Read the production columns
    # P = column 16
    # User said: P109 = Real Ton, P112 = Potensi Ton
    
    real_ton_col = 16  # Column P (for Ton under "Real")
    # But there might be sub-columns... let me read a range
    
    print(f"\n📊 F008A (Row {f008_row}) Production Data:")
    print(f"\nColumns O-T (15-20) around column P:")
    for col_idx in range(15, 21):
        col_letter = ws.cell(1, col_idx).column_letter
        value = ws.cell(f008_row, col_idx).value
        print(f"  {col_letter}{f008_row}: {value}")
    
    # Also check what the headers are in row above F008A
    header_row = f008_row - 1
    print(f"\nHeader row ({header_row}) for columns O-T:")
    for col_idx in range(15, 21):
        col_letter = ws.cell(1, col_idx).column_letter
        value = ws.cell(header_row, col_idx).value
        print(f"  {col_letter}{header_row}: {value}")

# Now find what column has "Ton" under "Real 2025"
print("\n🔎 Searching for '2025' and 'Real/Potensi' headers...")
for row_idx in range(1, 20):
    row_vals = [ws.cell(row_idx, c).value for c in range(1, min(180, ws.max_column))]
    if '2025' in str(row_vals):
        print(f"Row {row_idx}: Found 2025")
        print(f"  Values: {row_vals[100:120]}")

wb.close()
