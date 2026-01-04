"""
FINAL EXTRACTION - ALL 36 BLOCKS PRODUCTION DATA
Columns found:
- 12: Luas (Ha)
- 171: Realisasi Ton (2025)
- 174: Potensi Ton (2025)
"""
from openpyxl import load_workbook
import json

print("✅ EXTRACTING ALL 36 BLOCKS - REAL PRODUCTION DATA!")
print("="*70)

wb = load_workbook('poac_sim/data/input/data_gabungan.xlsx', data_only=True)
ws = wb.active

# Load our 36 blocks list
with open('data/output/all_blocks_data.json') as f:
    blocks_data = json.load(f)

our_blocks = sorted(blocks_data.keys())
print(f"Target: {len(our_blocks)} AME II blocks\n")

# Extract production data
production_results = {}

for block_code in our_blocks:
    # Find block row
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 1).value).strip() == block_code:
            luas_ha = ws.cell(row, 12).value
            real_ton_total = ws.cell(row, 171).value or 0
            pot_ton_total = ws.cell(row, 174).value or 0
            
            # Calculate ton/ha
            if luas_ha and luas_ha > 0:
                real_ton_ha = real_ton_total / luas_ha
                pot_ton_ha = pot_ton_total / luas_ha
                gap_ton_ha = real_ton_ha - pot_ton_ha
                gap_pct = (gap_ton_ha / pot_ton_ha) * 100 if pot_ton_ha > 0 else 0
            else:
                real_ton_ha = pot_ton_ha = gap_ton_ha = gap_pct = 0
            
            production_results[block_code] = {
                'luas_ha': round(luas_ha, 2) if luas_ha else 0,
                'realisasi_ton_total': round(real_ton_total, 2) if real_ton_total else 0,
                'potensi_ton_total': round(pot_ton_total, 2) if pot_ton_total else 0,
                'realisasi_ton_ha': round(real_ton_ha, 2),
                'potensi_ton_ha': round(pot_ton_ha, 2),
                'gap_ton_ha': round(gap_ton_ha, 2),
                'gap_pct': round(gap_pct, 1)
            }
            
            print(f"✅ {block_code}: {real_ton_ha:.2f} ton/ha (gap: {gap_pct:+.1f}%)")
            break

wb.close()

print(f"\n✅ Extracted {len(production_results)}/{len(our_blocks)} blocks")

# Save results
with open('data/output/all_36_blocks_production_data.json', 'w') as f:
    json.dump(production_results, f, indent=2)

print(f"\n✅ Saved to: all_36_blocks_production_data.json")

# Show summary stats
gaps = [v['gap_pct'] for v in production_results.values() if v['gap_pct'] != 0]
print(f"\n📊 SUMMARY:")
print(f"  Blocks with production data: {len([v for v in production_results.values() if v['realisasi_ton_ha'] > 0])}")
print(f"  Positive gap (surplus): {len([g for g in gaps if g > 0])}")
print(f"  Negative gap (deficit): {len([g for g in gaps if g < 0])}")
print(f"  Average gap: {sum(gaps)/len(gaps):.1f}% (if gaps exist)")
